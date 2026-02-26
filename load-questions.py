#!/usr/bin/env python3
"""Convert quiz-template.xlsx to questions.json"""

import json
import sys
from openpyxl import load_workbook

xlsx = sys.argv[1] if len(sys.argv) > 1 else 'quiz-template.xlsx'
wb = load_workbook(xlsx, data_only=True)
ws = wb.active

letter_map = {'A': 0, 'B': 1, 'C': 2, 'D': 3, '1': 0, '2': 1, '3': 2, '4': 3}
questions = []

for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
    question, opt_a, opt_b, opt_c, opt_d, correct = row
    if not question or not correct:
        break
    correct = str(correct).strip().upper()
    if correct not in letter_map:
        print(f"Warning: Skipping row — invalid correct answer '{correct}' (must be A/B/C/D or 1/2/3/4)")
        continue
    questions.append({
        'question': str(question).strip(),
        'options': [str(opt_a).strip(), str(opt_b).strip(), str(opt_c).strip(), str(opt_d).strip()],
        'correct': letter_map[correct]
    })

# Prompt for title
title = input(f'Quiz title [{len(questions)} questions loaded]: ').strip()
if not title:
    from datetime import datetime
    title = datetime.now().strftime('%B %Y') + ' Product Quiz'

output = {
    'title': title,
    'timePerQuestion': 15,
    'questions': questions
}

with open('questions.json', 'w') as f:
    json.dump(output, f, indent=2)

print(f'Saved {len(questions)} questions to questions.json')
print(f'Title: {title}')
